import csv
import io
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple, Union

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from services.product_context_service import _CURRENCY_ALIASES, _clean_name


@dataclass
class CatalogIssue:
    severity: str  # "error" | "warning"
    code: str
    row: Optional[int] = None
    sheet: Optional[str] = None
    field: Optional[str] = None
    message: str = ""

    def to_dict(self) -> Dict[str, Any]:
        return {
            "severity": self.severity,
            "code": self.code,
            "row": self.row,
            "sheet": self.sheet,
            "field": self.field,
            "message": self.message,
        }


@dataclass
class CatalogStats:
    total_rows_seen: int = 0
    accepted_records: int = 0
    rejected_rows: int = 0
    product_count: int = 0
    bundle_count: int = 0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "total_rows_seen": self.total_rows_seen,
            "accepted_records": self.accepted_records,
            "rejected_rows": self.rejected_rows,
            "product_count": self.product_count,
            "bundle_count": self.bundle_count,
        }


@dataclass
class CatalogParseResult:
    records: List[Dict[str, Any]] = field(default_factory=list)
    issues: List[Dict[str, Any]] = field(default_factory=list)
    stats: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "records": self.records,
            "issues": self.issues,
            "stats": self.stats,
        }


_HEADER_ALIASES: Dict[str, str] = {
    # record_type
    "record_type": "record_type",
    "type": "record_type",
    "نوع_السجل": "record_type",
    "نوع السجل": "record_type",
    "نوع_سجل": "record_type",
    "نوع سجل": "record_type",
    # sku
    "sku": "sku",
    "product_code": "sku",
    "code": "sku",
    "كود": "sku",
    "كود_المنتج": "sku",
    "كود المنتج": "sku",
    "رمز_المنتج": "sku",
    "رمز المنتج": "sku",
    # category
    "category": "category",
    "الفئة": "category",
    "التصنيف": "category",
    "فئة": "category",
    "تصنيف": "category",
    # name
    "name": "name",
    "product_name": "name",
    "اسم": "name",
    "الاسم": "name",
    "اسم_المنتج": "name",
    "اسم المنتج": "name",
    "المنتج": "name",
    # description
    "description": "description",
    "الوصف": "description",
    "وصف": "description",
    # components
    "components": "components_text",
    "المكونات": "components_text",
    "مكونات": "components_text",
    # description_or_components
    "description_or_components": "description_or_components",
    "الوصف_او_المكونات": "description_or_components",
    "الوصف أو المكونات": "description_or_components",
    "الوصف_أو_المكونات": "description_or_components",
    "الوصف/المكونات": "description_or_components",
    # price
    "price": "price",
    "unit_price": "price",
    "السعر": "price",
    "سعر": "price",
    "سعر_الوحدة": "price",
    "سعر الوحدة": "price",
    # currency
    "currency": "currency",
    "curr": "currency",
    "العملة": "currency",
    "عملة": "currency",
    # stock
    "stock": "stock",
    "inventory": "stock",
    "المخزون": "stock",
    "الكمية_المتاحة": "stock",
    "الكمية المتاحة": "stock",
    "مخزون": "stock",
    # colors
    "colors": "colors",
    "colours": "colors",
    "الألوان": "colors",
    "الالوان": "colors",
    "ألوان": "colors",
    "الوان": "colors",
    # warranty
    "warranty": "warranty",
    "الضمان": "warranty",
    "ضمان": "warranty",
    # installation
    "installation": "installation",
    "التركيب": "installation",
    "تركيب": "installation",
    # installation_fee
    "installation_fee": "installation_fee",
    "installation_price": "installation_fee",
    "سعر_التركيب": "installation_fee",
    "سعر التركيب": "installation_fee",
    "رسوم_التركيب": "installation_fee",
    "رسوم التركيب": "installation_fee",
    # aliases
    "aliases": "aliases",
    "alias": "aliases",
    "الأسماء_البديلة": "aliases",
    "الأسماء البديلة": "aliases",
    "أسماء_بديلة": "aliases",
    "أسماء بديلة": "aliases",
    "أسماء_بديله": "aliases",
    "أسماء بديله": "aliases",
    # notes
    "notes": "notes",
    "ملاحظات": "notes",
    "ملاحظة": "notes",
}


def _normalize_header(header: str) -> str:
    h = header.strip()
    key_norm = re.sub(r"\s+", " ", h)
    key_under = key_norm.replace(" ", "_")

    if h.casefold() in _HEADER_ALIASES:
        return _HEADER_ALIASES[h.casefold()]
    if key_norm.casefold() in _HEADER_ALIASES:
        return _HEADER_ALIASES[key_norm.casefold()]
    if key_under.casefold() in _HEADER_ALIASES:
        return _HEADER_ALIASES[key_under.casefold()]

    return h


def _parse_tier_header(header: str) -> Optional[Tuple[int, Optional[int]]]:
    raw = header.strip()

    m_range = re.search(r"(?:خصم|discount)[_\s]*(\d+)[_\s]*(?:الى|إلى|to|-|_)[_\s]*(\d+)", raw, re.IGNORECASE)
    if m_range:
        return (int(m_range.group(1)), int(m_range.group(2)))

    m_plus = re.search(r"(?:خصم|discount)[_\s]*(\d+)[_\s]*(?:فاكثر|فأكثر|واكثر|وأكثر|فما_فوق|plus|and_above|\+)", raw, re.IGNORECASE)
    if m_plus:
        return (int(m_plus.group(1)), None)

    m_standalone_range = re.search(r"^(\d+)[_\s]*(?:الى|إلى|to|-)[_\s]*(\d+)$", raw, re.IGNORECASE)
    if m_standalone_range:
        return (int(m_standalone_range.group(1)), int(m_standalone_range.group(2)))

    m_standalone_plus = re.search(r"^(\d+)[_\s]*(?:plus|\+|\bفاكثر|\bفأكثر)$", raw, re.IGNORECASE)
    if m_standalone_plus:
        return (int(m_standalone_plus.group(1)), None)

    return None


def _parse_number(val: Any) -> Tuple[Optional[float], bool]:
    if val is None:
        return None, False
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return (float(val), False) if val >= 0 else (None, True)

    text = str(val).strip()
    if not text:
        return None, False

    clean_text = text.replace(",", "").replace(" ", "")
    if re.match(r"^\d+(?:\.\d+)?$", clean_text):
        try:
            num = float(clean_text)
            return (num, False) if num >= 0 else (None, True)
        except ValueError:
            return None, True
    return None, True


def _parse_integer(val: Any) -> Tuple[Optional[int], bool]:
    if val is None:
        return None, False
    if isinstance(val, bool):
        return None, True
    if isinstance(val, int):
        return (val, False) if val >= 0 else (None, True)
    if isinstance(val, float):
        if val >= 0 and val.is_integer():
            return int(val), False
        return None, True

    text = str(val).strip()
    if not text:
        return None, False

    clean_text = text.replace(",", "").replace(" ", "")
    if re.match(r"^\d+$", clean_text):
        return int(clean_text), False
    if re.match(r"^\d+\.0+$", clean_text):
        return int(float(clean_text)), False

    return None, True


def _parse_discount_pct(val: Any) -> Tuple[Optional[float], bool]:
    if val is None:
        return None, False
    num, is_invalid = _parse_number(val)
    if is_invalid or num is None:
        text = str(val).strip()
        if not text:
            return None, False
        return None, True

    if 0.0 <= num <= 100.0:
        return num, False
    return None, True


def _normalize_currency(val: Any) -> Optional[str]:
    if val is None:
        return None
    text = str(val).strip()
    if not text:
        return None
    key = text.casefold()
    if key in _CURRENCY_ALIASES:
        return _CURRENCY_ALIASES[key]
    return text.upper()


def _parse_list_field(val: Any) -> List[str]:
    if val is None:
        return []
    text = str(val).strip()
    if not text:
        return []

    parts = re.split(r"[,،;]", text)
    cleaned = [_clean_name(part) for part in parts]
    non_empty = [p for p in cleaned if p]
    return list(dict.fromkeys(non_empty))


def _normalize_record_type(val: Any) -> Tuple[Optional[str], bool]:
    if val is None:
        return "product", False
    text = str(val).strip().casefold()
    if not text:
        return "product", False

    if text in ("product", "منتج"):
        return "product", False
    if text in ("bundle", "package", "باقة", "حزمة"):
        return "bundle", False

    return None, True


def _parse_row(
    row_dict: Dict[str, Any],
    row_num: int,
    sheet_name: Optional[str] = None
) -> Tuple[Optional[Dict[str, Any]], List[CatalogIssue]]:
    issues: List[CatalogIssue] = []

    mapped_values: Dict[str, Any] = {}
    extra_fields: Dict[str, str] = {}
    quantity_discounts_raw: List[Tuple[Tuple[int, Optional[int]], Any, str]] = []

    for raw_header, raw_val in row_dict.items():
        if raw_header is None:
            continue
        header_str = str(raw_header).strip()
        if not header_str:
            continue

        canonical_field = _normalize_header(header_str)
        tier = _parse_tier_header(header_str)

        if tier is not None:
            quantity_discounts_raw.append((tier, raw_val, header_str))
        elif canonical_field in (
            "record_type", "sku", "category", "name", "description",
            "components_text", "description_or_components", "price",
            "currency", "stock", "colors", "warranty", "installation",
            "installation_fee", "aliases", "notes"
        ):
            mapped_values[canonical_field] = raw_val
        else:
            val_str = str(raw_val).strip() if raw_val is not None else ""
            if val_str:
                extra_fields[header_str] = val_str

    # Record Type
    raw_rec_type = mapped_values.get("record_type")
    record_type, rec_type_invalid = _normalize_record_type(raw_rec_type)
    if rec_type_invalid:
        issues.append(
            CatalogIssue(
                severity="error",
                code="INVALID_RECORD_TYPE",
                row=row_num,
                sheet=sheet_name,
                field="record_type",
                message=f"Unrecognized record_type '{raw_rec_type}'.",
            )
        )
        return None, issues

    # Name (Required)
    raw_name = mapped_values.get("name")
    name = _clean_name(raw_name)
    if not name:
        issues.append(
            CatalogIssue(
                severity="error",
                code="MISSING_NAME",
                row=row_num,
                sheet=sheet_name,
                field="name",
                message="Required field 'name' is missing or empty.",
            )
        )
        return None, issues

    sku = _clean_name(mapped_values.get("sku")) or None
    category = _clean_name(mapped_values.get("category")) or None
    warranty = _clean_name(mapped_values.get("warranty")) or None
    installation = _clean_name(mapped_values.get("installation")) or None
    notes = _clean_name(mapped_values.get("notes")) or None

    desc_val = _clean_name(mapped_values.get("description")) or None
    comp_val = _clean_name(mapped_values.get("components_text")) or None
    desc_or_comp = _clean_name(mapped_values.get("description_or_components")) or None

    description = desc_val
    components_text = comp_val

    if desc_or_comp:
        if record_type == "bundle":
            if not components_text:
                components_text = desc_or_comp
        else:
            if not description:
                description = desc_or_comp

    price, price_invalid = _parse_number(mapped_values.get("price"))
    if price_invalid:
        issues.append(
            CatalogIssue(
                severity="warning",
                code="INVALID_PRICE",
                row=row_num,
                sheet=sheet_name,
                field="price",
                message=f"Price value '{mapped_values.get('price')}' is not a valid non-negative number.",
            )
        )

    currency = _normalize_currency(mapped_values.get("currency"))

    stock, stock_invalid = _parse_integer(mapped_values.get("stock"))
    if stock_invalid:
        issues.append(
            CatalogIssue(
                severity="warning",
                code="INVALID_STOCK",
                row=row_num,
                sheet=sheet_name,
                field="stock",
                message=f"Stock value '{mapped_values.get('stock')}' is not a valid non-negative integer.",
            )
        )

    colors = _parse_list_field(mapped_values.get("colors"))
    aliases = _parse_list_field(mapped_values.get("aliases"))

    inst_fee, inst_fee_invalid = _parse_number(mapped_values.get("installation_fee"))
    if inst_fee_invalid:
        issues.append(
            CatalogIssue(
                severity="warning",
                code="INVALID_INSTALLATION_FEE",
                row=row_num,
                sheet=sheet_name,
                field="installation_fee",
                message=f"Installation fee '{mapped_values.get('installation_fee')}' is not a valid non-negative number.",
            )
        )

    quantity_discounts: List[Dict[str, Any]] = []
    for tier, raw_disc_val, col_name in quantity_discounts_raw:
        disc_pct, disc_invalid = _parse_discount_pct(raw_disc_val)
        if disc_invalid:
            issues.append(
                CatalogIssue(
                    severity="warning",
                    code="INVALID_DISCOUNT",
                    row=row_num,
                    sheet=sheet_name,
                    field=col_name,
                    message=f"Discount percentage '{raw_disc_val}' in column '{col_name}' is not a valid number from 0 to 100.",
                )
            )
        elif disc_pct is not None and disc_pct > 0:
            quantity_discounts.append(
                {
                    "min_qty": tier[0],
                    "max_qty": tier[1],
                    "discount_pct": disc_pct,
                }
            )

    record = {
        "record_type": record_type,
        "sku": sku,
        "name": name,
        "category": category,
        "description": description,
        "price": price,
        "currency": currency,
        "stock": stock,
        "colors": colors,
        "warranty": warranty,
        "installation": installation,
        "installation_fee": inst_fee,
        "aliases": aliases,
        "quantity_discounts": quantity_discounts,
        "components_text": components_text,
        "notes": notes,
        "extra_fields": extra_fields,
    }

    return record, issues


def parse_catalog_csv(
    content: Union[bytes, str], sheet_name: Optional[str] = None
) -> CatalogParseResult:
    issues: List[Dict[str, Any]] = []
    records: List[Dict[str, Any]] = []
    stats = CatalogStats()

    if isinstance(content, bytes):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = content.decode("cp1256")
            except UnicodeDecodeError:
                issues.append(
                    CatalogIssue(
                        severity="error",
                        code="ENCODING_ERROR",
                        sheet=sheet_name,
                        message="Failed to decode CSV content with UTF-8 or CP1256.",
                    ).to_dict()
                )
                return CatalogParseResult(records=[], issues=issues, stats=stats.to_dict())
    else:
        text = content

    if not text.strip():
        issues.append(
            CatalogIssue(
                severity="error",
                code="EMPTY_FILE",
                sheet=sheet_name,
                message="CSV content is empty.",
            ).to_dict()
        )
        return CatalogParseResult(records=[], issues=issues, stats=stats.to_dict())

    first_line = ""
    for line in text.splitlines():
        if line.strip():
            first_line = line
            break

    delimiters = [",", ";", "\t"]
    counts = {d: first_line.count(d) for d in delimiters}
    best_delimiter = max(counts, key=counts.get) if counts else ","
    if counts[best_delimiter] == 0:
        best_delimiter = ","

    stream = io.StringIO(text)
    reader = csv.reader(stream, delimiter=best_delimiter)

    headers: Optional[List[str]] = None
    line_number = 0

    for row_cells in reader:
        line_number += 1
        if not row_cells or not any(str(c).strip() for c in row_cells):
            continue

        if headers is None:
            headers = [str(c).strip() for c in row_cells]
            recognized_count = 0
            for h in headers:
                norm_h = _normalize_header(h)
                tier_h = _parse_tier_header(h)
                if norm_h in _HEADER_ALIASES.values() or tier_h is not None:
                    recognized_count += 1

            if recognized_count == 0:
                if counts[best_delimiter] == 0:
                    issues.append(
                        CatalogIssue(
                            severity="error",
                            code="UNSUPPORTED_DELIMITER",
                            sheet=sheet_name,
                            message="CSV header row does not contain a supported delimiter (comma, semicolon, tab) or recognized catalog header.",
                        ).to_dict()
                    )
                else:
                    issues.append(
                        CatalogIssue(
                            severity="warning",
                            code="NO_CATALOG_HEADER",
                            sheet=sheet_name,
                            message="CSV header row does not contain recognizable catalog headers.",
                        ).to_dict()
                    )
                return CatalogParseResult(records=[], issues=issues, stats=stats.to_dict())
            continue

        stats.total_rows_seen += 1
        row_dict = {
            headers[i]: row_cells[i]
            for i in range(min(len(headers), len(row_cells)))
        }

        record, row_issues = _parse_row(row_dict, row_num=line_number, sheet_name=sheet_name)
        for iss in row_issues:
            issues.append(iss.to_dict())

        if record is not None:
            records.append(record)
            stats.accepted_records += 1
            if record["record_type"] == "product":
                stats.product_count += 1
            elif record["record_type"] == "bundle":
                stats.bundle_count += 1
        else:
            stats.rejected_rows += 1

    return CatalogParseResult(records=records, issues=issues, stats=stats.to_dict())


def parse_catalog_xlsx(content: bytes) -> CatalogParseResult:
    issues: List[Dict[str, Any]] = []
    records: List[Dict[str, Any]] = []
    stats = CatalogStats()

    if not HAS_OPENPYXL:
        issues.append(
            CatalogIssue(
                severity="error",
                code="DEPENDENCY_MISSING",
                message="openpyxl package is required for XLSX parsing.",
            ).to_dict()
        )
        return CatalogParseResult(records=[], issues=issues, stats=stats.to_dict())

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        issues.append(
            CatalogIssue(
                severity="error",
                code="XLSX_LOAD_ERROR",
                message=f"Failed to open XLSX workbook: {str(e)}",
            ).to_dict()
        )
        return CatalogParseResult(records=[], issues=issues, stats=stats.to_dict())

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        header_row_idx = None
        headers = None

        for idx, row in enumerate(all_rows[:20]):
            if not row:
                continue
            str_cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if not str_cells:
                continue

            recognized_count = 0
            for cell in str_cells:
                norm_h = _normalize_header(cell)
                tier_h = _parse_tier_header(cell)
                if norm_h in _HEADER_ALIASES.values() or tier_h is not None:
                    recognized_count += 1

            if recognized_count >= 1:
                header_row_idx = idx
                headers = [str(c).strip() if c is not None else "" for c in row]
                break

        if header_row_idx is None or headers is None:
            has_some_content = any(any(c is not None for c in r) for r in all_rows)
            if has_some_content:
                issues.append(
                    CatalogIssue(
                        severity="warning",
                        code="NO_CATALOG_HEADER",
                        sheet=sheet_name,
                        message=f"Worksheet '{sheet_name}' does not contain recognizable catalog headers and was skipped.",
                    ).to_dict()
                )
            continue

        for idx in range(header_row_idx + 1, len(all_rows)):
            row = all_rows[idx]
            line_number = idx + 1
            if not row or not any(c is not None and str(c).strip() for c in row):
                continue

            stats.total_rows_seen += 1
            row_dict = {
                headers[i]: row[i]
                for i in range(min(len(headers), len(row)))
                if i < len(headers) and headers[i]
            }

            record, row_issues = _parse_row(row_dict, row_num=line_number, sheet_name=sheet_name)
            for iss in row_issues:
                issues.append(iss.to_dict())

            if record is not None:
                records.append(record)
                stats.accepted_records += 1
                if record["record_type"] == "product":
                    stats.product_count += 1
                elif record["record_type"] == "bundle":
                    stats.bundle_count += 1
            else:
                stats.rejected_rows += 1

    return CatalogParseResult(records=records, issues=issues, stats=stats.to_dict())


def parse_catalog_bytes(content: bytes, filename: str) -> CatalogParseResult:
    fn_lower = filename.lower()
    if fn_lower.endswith(".xlsx"):
        return parse_catalog_xlsx(content)
    elif fn_lower.endswith(".csv"):
        return parse_catalog_csv(content)
    else:
        return parse_catalog_csv(content)
